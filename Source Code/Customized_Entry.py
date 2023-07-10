#importing openpxyl to read excel files
from openpyxl import load_workbook
print('MADE BY KAUSHIL NAGRALE\n')
print('*************EASY PWD (https://ams.emahapwd.com/ams/login.do entry) *************\n')
print('*************NOTE ENTER DATA FROM FIRST ROW *************\n')
f= open("javaSc.txt","w+")
print('Keep the File inside the same directory as this .py file\n')

#Taking the details of Excel file as inputs
path_excel=input('1. Enter the Path (Excel in file in which data is present)(.xlsx ONLY)\n')
work_sheet_name='auto'
y=input('2. Enter the no of rows\n')
count=1
count=int(count)

#Data-based adding of rows
y=int(y)
while count < y:
     f.write('addRows();\n')
     count=count+1


#counting and iteration variable
x = 1

#Creating list to store the excel data
re,nu,le,br,de=[],[],[],[],[]
wb = load_workbook(path_excel)
ws = wb.get_sheet_by_name(work_sheet_name)
ha = 'A'
column = ws[ha]
re = [column[a].value for a in range(len(column))]
hb = 'B'
column = ws[hb]  # Column
nu = [column[b].value for b in range(len(column))]
hc = 'C'
column = ws[hc]  # Column
le = [column[c].value for c in range(len(column))]
hd = 'D'
column = ws[hd]  # Column
br = [column[d].value for d in range(len(column))]
he = 'E' # Work Sheet
column = ws[he]  # Column
de = [column[e].value for e in range(len(column))]

#Loop for to print the Javascript Command to insert data in rows
try:
    while x <= y:
      row = ('''document.getElementById('remarks%d').value = '%s';\n''' %(x,re[x-1]))
      if re[x-1]  != None:
           f.write(row)
      row = ('''document.getElementById('num%d').value = '%s';\n''' %(x,nu[x-1]))
      f.write(row)
      row = ('''document.getElementById('len%d').value = '%s';\n''' %(x,le[x-1]))
      f.write(row)
      row = ('''document.getElementById('bre%d').value = '%s';\n'''%(x,br[x-1]))
      f.write(row)
      row = ('''document.getElementById('dep%d').value = '%s';\n''' %(x,de[x-1]))
      f.write(row)
      row = ('''document.getElementById('hm%d').checked  = true;\n''' %x)
      f.write(row)
      x=x+1
except:
    print("Not enough data")
  
#Loop for to print the Javascript Command to insert data in rows
count=1
while count <= y:
    f.write('calculateQty(%d);\n'%count)
    count=count+1
    
print('Sucessful!! \n copy the contents of the file of .txt to javascript console')
