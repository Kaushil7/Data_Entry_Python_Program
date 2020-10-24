#importing openpxyl to read excel files
from openpyxl import load_workbook
print('*************NOTE ENTER DATA FROM FIRST ROW *************\nMADE BY KAUSHIL NAGRALE\n')
print('*************EASY PWD (https://ams.emahapwd.com/ams/login.do entry) *************\n')
f= open("COPY_THIS_FILE_TO_WEBSITE_JS.txt","w+")


y=input('*************Enter the no of rows*************\n')
count=int(1)


# data based adding of rows
y=int(y)
while count < y:
     f.write('addRows();\n')
     count=count+1


#Creating list to store the excel data
re,nu,le,br,de=[],[],[],[],[]
wb = load_workbook("auto.xlsx")
ws = wb.get_sheet_by_name("auto")
column = ws['A']
re = [column[a].value for a in range(len(column))]
column = ws['B']
nu = [column[b].value for b in range(len(column))]
column = ws['C']
le = [column[c].value for c in range(len(column))]
column = ws['D']
br = [column[d].value for d in range(len(column))]
column = ws['E']
de = [column[e].value for e in range(len(column))]


#counting and iteration variable
x = 1


#Loop for to print the Javascript Command to insert data in rows
try:
    while x <= y:
      row = ('''document.getElementById('remarks%d').value = '%s';\n''' %(x,re[x-1]))
      if re[x-1]  != None:
           f.write(row)
      row = ('''document.getElementById('num%d').value = '%s';\n''' %(x,nu[x-1]))
      f.write(row)
      row = ('''document.getElementById('len%d').value = '%s';\n''' %(x,le[x-1]))
      f.write(row)
      row = ('''document.getElementById('bre%d').value = '%s';\n'''%(x,br[x-1]))
      f.write(row)
      row = ('''document.getElementById('dep%d').value = '%s';\n''' %(x,de[x-1]))
      f.write(row)
      row = ('''document.getElementById('hm%d').checked  = true;\n''' %x)
      f.write(row)
      x=x+1
except:
    print("Not enough data")


#Loop to print the Javascript command to calculate the quantity on the website
count=1
while count <=y:
    f.write('calculateQty(%d);\n'%count)
    count=count+1
