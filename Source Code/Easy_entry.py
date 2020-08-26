from openpyxl import load_workbook
print('*************NOTE ENTER DATA FROM FIRST ROW *************\nMADE BY KAUSHIL NAGRALE\n')
print('*************EASY PWD (https://ams.emahapwd.com/ams/login.do entry) *************\n')
f= open("COPY_THIS_FILE_TO_WEBSITE_DATA.txt","w+")


y=input('*************Enter the no of rows*************\n')
count=1
count=int(count)
y=int(y)
while count < y:
     f.write('addRows();\n')
     count=count+1

cou=y+1
x = 1

re=[]
wb = load_workbook("auto.xlsx")
ws = wb.get_sheet_by_name("auto")
column = ws['A']
re = [column[a].value for a in range(len(column))]

nu=[]
wb = load_workbook("auto.xlsx")
ws = wb.get_sheet_by_name("auto")

column = ws['B']
nu = [column[b].value for b in range(len(column))]

le=[]
wb = load_workbook("auto.xlsx")
ws = wb.get_sheet_by_name("auto")

column = ws['C']
le = [column[c].value for c in range(len(column))]

br=[]
wb = load_workbook("auto.xlsx")
ws = wb.get_sheet_by_name("auto")

column = ws['D']
br = [column[d].value for d in range(len(column))]

de=[]
wb = load_workbook("auto.xlsx")
ws = wb.get_sheet_by_name("auto")
column = ws['E']
de = [column[e].value for e in range(len(column))]

cou=int(cou)
try:
    while x < cou:
      row = ('''document.getElementById('remarks%d').value = '%s';\n''' %(x,re[x-1]))
      if re[x-1]  != None:
           f.write(row)
      row = ('''document.getElementById('num%d').value = '%s';\n''' %(x,nu[x-1]))
      f.write(row)
      row = ('''document.getElementById('len%d').value = '%s';\n''' %(x,le[x-1]))
      f.write(row)
      row = ('''document.getElementById('bre%d').value = '%s';\n'''%(x,br[x-1]))
      f.write(row)
      row = ('''document.getElementById('dep%d').value = '%s';\n''' %(x,de[x-1]))
      f.write(row)
      row = ('''document.getElementById('hm%d').checked  = true;\n''' %x)
      f.write(row)
      x=x+1
except:
    print("Not enough data")

count=1
while count < cou:
    f.write('calculateQty(%d);\n'%count)
    count=count+1
