#importing openpxyl to read excel files
from openpyxl import load_workbook
import os

# get the current working directory
current_working_directory = os.getcwd()

print('MADE BY KAUSHIL NAGRALE\n')
print('*************EASY PWD (https://ams.emahapwd.com/ams/login.do entry) *************\n')
print('*************NOTE ENTER DATA FROM FIRST ROW *************\n')
f= open("javaSc.txt","w+")
print('Keep the File inside the same directory as this .py file\n')

#Taking the details of Excel file as inputs
path_excel=current_working_directory + '/Source Code/auto.xlsx' 
path_excel= path_excel.replace('\\','/')
work_sheet_name='auto'
rowss=int(input('Enter the no of rows\n'))

#Data-based adding of rows
for x in range(2,rowss+1):
    f.write('addRows();\n')

#counting and iteration variable
x = 1

#Creating list to store the excel data
re,nu,le,br,de=[],[],[],[],[]
wb = load_workbook(path_excel)
ws = wb.get_sheet_by_name(work_sheet_name)
ha = 'A'
column = ws[ha]
re = [column[a].value for a in range(len(column))]
hb = 'B'
column = ws[hb]  # Column
nu = [column[b].value for b in range(len(column))]
hc = 'C'
column = ws[hc]  # Column
le = [column[c].value for c in range(len(column))]
hd = 'D'
column = ws[hd]  # Column
br = [column[d].value for d in range(len(column))]
he = 'E' # Work Sheet
column = ws[he]  # Column
de = [column[e].value for e in range(len(column))]

#Loop for to print the Javascript Command to insert data in rows
try:
    for x in range(1,rowss+1):
      row = ('''document.getElementById('remarks%d').value = '%s';\n''' %(x,re[x-1]))
      if re[x-1]  != None:
           f.write(row)
      row = ('''document.getElementById('num%d').value = '%s';\n''' %(x,nu[x-1]))
      f.write(row)
      row = ('''document.getElementById('len%d').value = '%s';\n''' %(x,le[x-1]))
      f.write(row)
      row = ('''document.getElementById('bre%d').value = '%s';\n'''%(x,br[x-1]))
      f.write(row)
      row = ('''document.getElementById('dep%d').value = '%s';\n''' %(x,de[x-1]))
      f.write(row)
      row = ('''document.getElementById('hm%d').checked  = true;\n''' %x)
      f.write(row)      
except:
    print("Not enough data")
  
#Loop for to print the Javascript Command to insert data in rows
for x in range(1,rowss+1):
    f.write('calculateQty(%d);\n'%x)
    
print('Sucessful!! \n copy the contents of the file of .txt to javascript console')
